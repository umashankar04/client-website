import os
import threading
from openpyxl import Workbook, load_workbook

# Thread lock to prevent concurrent writes to Excel files
excel_lock = threading.Lock()

def init_excel_file(file_path, headers):
    """
    Initializes an Excel file with headers if it does not exist.
    """
    # Ensure the parent directory exists
    dir_name = os.path.dirname(file_path)
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

    # Use the lock to check and create file safely
    with excel_lock:
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Write headers in the first row
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            wb.save(file_path)
            wb.close()
            return True
    return False

def read_sheet(file_path):
    """
    Reads an Excel sheet and returns a list of dictionaries representing the rows.
    Each dictionary keys will match the headers in the first row.
    """
    if not os.path.exists(file_path):
        return []

    with excel_lock:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Read headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        rows = []
        # Iterate through remaining rows
        for r in range(2, ws.max_row + 1):
            row_data = {}
            row_is_empty = True
            for c_idx, header in enumerate(headers, 1):
                val = ws.cell(row=r, column=c_idx).value
                if val is not None:
                    row_is_empty = False
                row_data[header] = val
            
            # Add row if it's not completely blank
            if not row_is_empty:
                rows.append(row_data)
                
        wb.close()
        return rows

def write_sheet(file_path, data, headers):
    """
    Overwrites the Excel file with the provided data (list of dictionaries).
    """
    with excel_lock:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            
        # Write data rows
        for row_idx, row_dict in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                val = row_dict.get(header, None)
                # Convert list/dict to string if accidentally passed
                if isinstance(val, (list, dict)):
                    val = str(val)
                ws.cell(row=row_idx, column=col_idx, value=val)
                
        wb.save(file_path)
        wb.close()
        return True
