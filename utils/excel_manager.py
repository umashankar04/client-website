"""
utils/excel_manager.py
─────────────────────
Simple helper functions to read and write Excel files.
Each Excel file is treated like a database table.
Rows are stored as Python dictionaries for easy access.
"""

import os
import threading
from openpyxl import Workbook, load_workbook

# This lock prevents two threads from writing at the same time
_lock = threading.Lock()


def ensure_file(filepath, headers):
    """
    Create an Excel file with header row if it does not exist yet.
    filepath  – full path to the .xlsx file
    headers   – list of column names  e.g. ["ID", "Name", "Price"]
    """
    # Make the parent folder if needed
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if not os.path.exists(filepath):
        with _lock:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(headers)   # Write header row
            wb.save(filepath)


def read_rows(filepath):
    """
    Read all rows from an Excel file.
    Returns a list of dictionaries.
    Example: [{"ID": "C001", "Name": "John"}, ...]
    Returns empty list if file does not exist.
    """
    if not os.path.exists(filepath):
        return []

    with _lock:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active

        # First row is always the header
        headers = [cell.value for cell in ws[1]]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip completely empty rows
            if all(v is None for v in row):
                continue
            # Build dictionary from header + value pairs
            row_dict = {}
            for h, v in zip(headers, row):
                row_dict[h] = v
            rows.append(row_dict)

        wb.close()
        return rows


def write_rows(filepath, rows, headers):
    """
    Overwrite the entire Excel file with new data.
    rows    – list of dictionaries
    headers – list of column names (defines column order)
    """
    with _lock:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Write header row
        ws.append(headers)

        # Write each data row
        for row_dict in rows:
            row_values = [row_dict.get(h) for h in headers]
            ws.append(row_values)

        wb.save(filepath)
        wb.close()


def append_row(filepath, row_dict, headers):
    """
    Add one new row to an existing Excel file.
    Faster than rewriting the whole file.
    """
    with _lock:
        if not os.path.exists(filepath):
            # File missing – create with headers first
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(headers)
        else:
            wb = load_workbook(filepath)
            ws = wb.active

        row_values = [row_dict.get(h) for h in headers]
        ws.append(row_values)
        wb.save(filepath)
        wb.close()
